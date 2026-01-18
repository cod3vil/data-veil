"""
Property-based tests for File Exporter module

Feature: data-desensitization-platform
"""

import pytest
from hypothesis import given, strategies as st, settings, assume
from hypothesis import HealthCheck
import tempfile
import os
from datetime import datetime
from io import BytesIO

from app.file_exporter import FileExporter, FileExportError

# Import document reading libraries for verification
from docx import Document
from openpyxl import load_workbook


# Hypothesis strategies for generating test data
@st.composite
def text_content(draw):
    """Generate text content with mixed Chinese and English"""
    chinese_chars = st.characters(
        whitelist_categories=('Lo',),
        min_codepoint=0x4E00,
        max_codepoint=0x9FFF
    )
    ascii_chars = st.characters(min_codepoint=32, max_codepoint=126)
    
    text = draw(st.text(
        alphabet=st.one_of(chinese_chars, ascii_chars),
        min_size=10,
        max_size=500
    ))
    
    assume(text.strip())
    return text


@st.composite
def metadata_dict(draw):
    """Generate metadata dictionary"""
    metadata = {}
    
    # Optional fields
    if draw(st.booleans()):
        metadata['title'] = draw(st.text(min_size=1, max_size=50))
    if draw(st.booleans()):
        metadata['author'] = draw(st.text(min_size=1, max_size=50))
    if draw(st.booleans()):
        metadata['subject'] = draw(st.text(min_size=1, max_size=50))
    
    return metadata


# Feature: data-desensitization-platform, Property 16: Format-preserving Export
@given(
    content=text_content(),
    metadata=metadata_dict()
)
@settings(
    max_examples=100,
    deadline=None,
    suppress_health_check=[HealthCheck.function_scoped_fixture]
)
@pytest.mark.property_test
def test_txt_export_format_preservation(content, metadata):
    """
    Property 16: Format-preserving Export
    Validates: Requirements 7.1, 7.6
    
    For any desensitized document with TXT as original format,
    when exported with default settings, the output format should be TXT
    and content should be preserved.
    """
    exporter = FileExporter()
    
    # Export to TXT (default for TXT input)
    result = exporter.export(
        content=content,
        original_format='txt',
        output_format='txt',
        metadata=metadata
    )
    
    # Verify result is bytes
    assert isinstance(result, bytes)
    
    # Verify content is preserved
    decoded_content = result.decode('utf-8')
    assert decoded_content == content


@given(
    content=text_content(),
    metadata=metadata_dict()
)
@settings(
    max_examples=100,
    deadline=None,
    suppress_health_check=[HealthCheck.function_scoped_fixture]
)
@pytest.mark.property_test
def test_docx_export_format_preservation(content, metadata):
    """
    Property 16: Format-preserving Export
    Validates: Requirements 7.1, 7.4
    
    For any desensitized document with DOCX as original format,
    when exported with default settings, the output format should be DOCX
    and content should be preserved.
    """
    exporter = FileExporter()
    
    # Export to DOCX (default for DOCX input)
    result = exporter.export(
        content=content,
        original_format='docx',
        output_format='docx',
        metadata=metadata
    )
    
    # Verify result is bytes
    assert isinstance(result, bytes)
    assert len(result) > 0
    
    # Verify it's a valid DOCX by reading it back
    buffer = BytesIO(result)
    doc = Document(buffer)
    
    # Extract text from document
    extracted_text = '\n'.join([p.text for p in doc.paragraphs if p.text.strip()])
    
    # Verify content is preserved (should contain original content)
    assert content in extracted_text or extracted_text in content
    
    # Verify metadata if provided
    if metadata.get('title'):
        # Metadata should be sanitized (control characters removed)
        sanitized_title = ''.join(char for char in metadata['title'] if ord(char) >= 0x20 or char in '\t\n\r')
        assert doc.core_properties.title == sanitized_title
    if metadata.get('author'):
        sanitized_author = ''.join(char for char in metadata['author'] if ord(char) >= 0x20 or char in '\t\n\r')
        assert doc.core_properties.author == sanitized_author


@given(
    content=text_content(),
    metadata=metadata_dict()
)
@settings(
    max_examples=100,
    deadline=None,
    suppress_health_check=[HealthCheck.function_scoped_fixture]
)
@pytest.mark.property_test
def test_xlsx_export_format_preservation(content, metadata):
    """
    Property 16: Format-preserving Export
    Validates: Requirements 7.1, 7.5
    
    For any desensitized document with XLSX as original format,
    when exported with default settings, the output format should be XLSX
    and content should be preserved.
    """
    exporter = FileExporter()
    
    # Export to XLSX (default for XLSX input)
    result = exporter.export(
        content=content,
        original_format='xlsx',
        output_format='xlsx',
        metadata=metadata
    )
    
    # Verify result is bytes
    assert isinstance(result, bytes)
    assert len(result) > 0
    
    # Verify it's a valid XLSX by reading it back
    buffer = BytesIO(result)
    wb = load_workbook(buffer)
    
    # Extract text from workbook
    extracted_text = []
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        for row in sheet.iter_rows(values_only=True):
            row_values = [str(cell) for cell in row if cell is not None and str(cell).strip()]
            if row_values:
                extracted_text.extend(row_values)
    
    full_extracted = ' '.join(extracted_text)
    
    # Verify content is preserved (should contain original content)
    assert content in full_extracted or any(word in full_extracted for word in content.split()[:5])
    
    wb.close()


# Feature: data-desensitization-platform, Property 17: Markdown Export Support
@given(
    content=text_content(),
    original_format=st.sampled_from(['pdf', 'docx', 'xlsx', 'txt']),
    metadata=metadata_dict()
)
@settings(
    max_examples=100,
    deadline=None,
    suppress_health_check=[HealthCheck.function_scoped_fixture]
)
@pytest.mark.property_test
def test_markdown_export_support(content, original_format, metadata):
    """
    Property 17: Markdown Export Support
    Validates: Requirements 7.2, 7.7
    
    For any desensitized document of any input format,
    when exported with MD format selected, the output should be
    a valid Markdown file.
    """
    exporter = FileExporter()
    
    # Export to MD regardless of original format
    result = exporter.export(
        content=content,
        original_format=original_format,
        output_format='md',
        metadata=metadata
    )
    
    # Verify result is bytes
    assert isinstance(result, bytes)
    
    # Verify content can be decoded as UTF-8
    decoded_content = result.decode('utf-8')
    assert len(decoded_content) > 0
    
    # Verify original content is preserved in markdown
    assert content in decoded_content or any(word in decoded_content for word in content.split()[:5])
    
    # Verify markdown metadata header if metadata provided
    if metadata:
        assert '---' in decoded_content  # YAML frontmatter
        assert 'desensitized: true' in decoded_content
        
        if metadata.get('title'):
            # Title should appear in frontmatter or as heading
            assert metadata['title'] in decoded_content


@given(
    content=text_content(),
    metadata=metadata_dict()
)
@settings(
    max_examples=100,
    deadline=None,
    suppress_health_check=[HealthCheck.function_scoped_fixture]
)
@pytest.mark.property_test
def test_markdown_export_with_tables(content, metadata):
    """
    Property 17: Markdown Export Support (with tables)
    Validates: Requirements 7.2, 7.7
    
    For any content with table-like structure (pipe-separated),
    when exported to MD, tables should be formatted correctly.
    """
    exporter = FileExporter()
    
    # Create content with table structure
    table_content = f"{content}\nColumn1 | Column2 | Column3\nValue1 | Value2 | Value3"
    
    # Export to MD
    result = exporter.export(
        content=table_content,
        original_format='txt',
        output_format='md',
        metadata=metadata
    )
    
    decoded_content = result.decode('utf-8')
    
    # Verify table formatting
    assert '|' in decoded_content
    # Markdown tables should have separator row with ---
    assert '---' in decoded_content


# Feature: data-desensitization-platform, Property 18: Export Filename Convention
@given(
    original_filename=st.text(
        alphabet=st.characters(min_codepoint=97, max_codepoint=122),
        min_size=5,
        max_size=20
    ),
    output_format=st.sampled_from(['txt', 'md', 'docx', 'xlsx']),
    timestamp=st.datetimes(
        min_value=datetime(2020, 1, 1),
        max_value=datetime(2030, 12, 31)
    )
)
@settings(max_examples=100, deadline=None)
@pytest.mark.property_test
def test_export_filename_convention(original_filename, output_format, timestamp):
    """
    Property 18: Export Filename Convention
    Validates: Requirements 7.8
    
    For any exported desensitized file, the filename should include
    the original filename and a timestamp in a consistent format.
    """
    exporter = FileExporter()
    
    # Add extension to original filename
    original_with_ext = f"{original_filename}.pdf"
    
    # Generate filename
    result_filename = exporter.generate_filename(
        original_filename=original_with_ext,
        output_format=output_format,
        timestamp=timestamp
    )
    
    # Verify filename structure
    assert isinstance(result_filename, str)
    
    # Should contain original filename (without extension)
    assert original_filename in result_filename
    
    # Should contain "desensitized"
    assert 'desensitized' in result_filename
    
    # Should contain timestamp in format YYYYMMDD_HHMMSS
    time_str = timestamp.strftime('%Y%m%d_%H%M%S')
    assert time_str in result_filename
    
    # Should end with correct extension
    assert result_filename.endswith(f'.{output_format}')
    
    # Should follow pattern: original_desensitized_YYYYMMDD_HHMMSS.ext
    expected_pattern = f"{original_filename}_desensitized_{time_str}.{output_format}"
    assert result_filename == expected_pattern


@given(
    original_filename=st.text(
        alphabet=st.characters(min_codepoint=97, max_codepoint=122),
        min_size=5,
        max_size=20
    ),
    output_format=st.sampled_from(['txt', 'md', 'docx', 'xlsx'])
)
@settings(max_examples=100, deadline=None)
@pytest.mark.property_test
def test_export_filename_without_timestamp(original_filename, output_format):
    """
    Property 18: Export Filename Convention (default timestamp)
    Validates: Requirements 7.8
    
    For any exported file without explicit timestamp,
    the filename should use current time.
    """
    exporter = FileExporter()
    
    # Add extension to original filename
    original_with_ext = f"{original_filename}.pdf"
    
    # Generate filename without timestamp (should use current time)
    result_filename = exporter.generate_filename(
        original_filename=original_with_ext,
        output_format=output_format
    )
    
    # Verify filename structure
    assert isinstance(result_filename, str)
    assert original_filename in result_filename
    assert 'desensitized' in result_filename
    assert result_filename.endswith(f'.{output_format}')
    
    # Should contain a timestamp (8 digits for date + underscore + 6 digits for time)
    # Pattern: YYYYMMDD_HHMMSS
    import re
    timestamp_pattern = r'\d{8}_\d{6}'
    assert re.search(timestamp_pattern, result_filename) is not None
